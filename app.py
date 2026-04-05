from __future__ import annotations
import io
import logging
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from data_loader import LoadResult, finalize_binding, load_file
from processor import apply_na_fill, process
from pricing_engine import CACHE_METADATA, DEFAULT_REGION, REGION_LABELS, SUPPORTED_REGIONS, cache_age_days, cache_is_stale
logging.basicConfig(level=logging.WARNING)
log = logging.getLogger(__name__)
st.set_page_config(page_title='FinOps Optimizer', page_icon='💰', layout='wide', initial_sidebar_state='collapsed')
st.markdown('\n<style>\n:root {\n  --fin-bg: #f4f5f7;\n  --fin-card: #ffffff;\n  --fin-text: #1a1d23;\n  --fin-muted: #6b7280;\n  --fin-border: #e2e5eb;\n  --fin-green: #00955c;\n  --fin-amber: #d97706;\n  --fin-red: #dc2626;\n  --fin-header-fg: #f0f2f5;\n  --fin-header-bg: #1a1d23;\n}\n@media (prefers-color-scheme: dark) {\n  :root {\n    --fin-bg: #0e1117;\n    --fin-card: #1e2128;\n    --fin-text: #f3f4f6;\n    --fin-muted: #9ca3af;\n    --fin-border: #374151;\n    --fin-header-fg: #f9fafb;\n    --fin-header-bg: #111827;\n  }\n}\nhtml, body {\n  font-family: system-ui, -apple-system, "Segoe UI", Roboto, sans-serif;\n}\n.stApp {\n  font-family: system-ui, -apple-system, "Segoe UI", Roboto, sans-serif;\n}\n@media (prefers-color-scheme: dark) {\n  .metric { background: var(--fin-card); border-color: var(--fin-border); }\n  .metric-label { color: var(--fin-muted); }\n  .metric-value { color: var(--fin-text); }\n}\n.hdr {\n  background: var(--fin-header-bg);\n  color: var(--fin-header-fg);\n  padding: 1.2rem 1.5rem;\n  border-radius: 8px;\n  margin-bottom: 1rem;\n}\n.hdr h1 { margin: 0; font-size: 1.35rem; font-weight: 600; }\n.hdr p { margin: 0.35rem 0 0; font-size: 0.8rem; opacity: 0.85; }\n.sec { font-size: 0.7rem; font-weight: 600; color: var(--fin-muted);\n       letter-spacing: 0.05em; text-transform: uppercase; margin-bottom: 0.35rem; }\n.metric-row { display: flex; gap: 0.75rem; flex-wrap: wrap; margin-bottom: 0.75rem; }\n.metric {\n  flex: 1 1 120px;\n  background: var(--fin-card);\n  border: 1px solid var(--fin-border);\n  border-radius: 8px;\n  padding: 0.85rem 1rem;\n}\n.metric-label { font-size: 0.65rem; font-weight: 600; color: var(--fin-muted); }\n.metric-value { font-size: 1.35rem; font-weight: 600; color: var(--fin-text); margin-top: 0.15rem; }\n.box-ok, .box-warn, .box-err {\n  padding: 0.65rem 1rem;\n  border-radius: 0 6px 6px 0;\n  font-size: 0.82rem;\n  margin: 0.4rem 0;\n}\n.box-ok { background: rgba(34, 197, 94, 0.12); border-left: 3px solid #22c55e; color: var(--fin-text); }\n.box-warn { background: rgba(245, 158, 11, 0.12); border-left: 3px solid #f59e0b; color: var(--fin-text); }\n.box-err { background: rgba(239, 68, 68, 0.12); border-left: 3px solid #ef4444; color: var(--fin-text); }\n#MainMenu, footer { visibility: hidden; }\n.block-container { padding-top: 1rem; max-width: 1480px; }\n[data-testid="stDataFrame"] { max-height: 520px; overflow: auto !important; }\n</style>\n', unsafe_allow_html=True)

def _savings_numeric(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        if v.strip() == 'No Savings':
            return 0.0
        try:
            return float(v.replace('%', ''))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def build_excel(df: pd.DataFrame, region_label: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Recommendations')
        wb = writer.book
        ws = writer.sheets['Recommendations']
        thin = Side(style='thin', color='888888')
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill('solid', fgColor='1A1D23')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for cidx in range(1, ws.max_column + 1):
            c = ws.cell(row=1, column=cidx)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sav_cols = [c for c in df.columns if 'Savings %' in c]
        price_cols = [c for c in df.columns if 'Cost ($)' in c]
        green_fill = PatternFill('solid', fgColor='D1FAE5')
        amber_fill = PatternFill('solid', fgColor='FEF3C7')
        red_fill = PatternFill('solid', fgColor='FEE2E2')
        col_list = list(df.columns)
        for ridx in range(2, ws.max_row + 1):
            for cidx in range(1, ws.max_column + 1):
                cell = ws.cell(row=ridx, column=cidx)
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(size=10)
            for name in sav_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                sc = ws.cell(row=ridx, column=ci)
                val = sc.value
                nv = _savings_numeric(val)
                if nv is None:
                    continue
                if nv >= 20:
                    sc.fill = green_fill
                elif nv > 0:
                    sc.fill = amber_fill
                else:
                    sc.fill = red_fill
            for name in price_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                ws.cell(row=ridx, column=ci).number_format = '$#,##0.0000'
        for col_cells in ws.columns:
            w = max((len(str(c.value or '')) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w + 2, 48)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws_m = wb.create_sheet('Metadata')
        ws_m.append(['Field', 'Value'])
        ws_m.append(['Generated at', datetime.now().strftime('%Y-%m-%d %H:%M')])
        ws_m.append(['Pricing region', region_label])
        ws_m.append(['Pricing vintage', CACHE_METADATA['last_updated'].strftime('%Y-%m-%d')])
        ws_m.append(['Rows', len(df)])
    return buf.getvalue()

def _savings_for_kpi(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str) and v.strip() == 'No Savings':
        return None
    return _savings_numeric(v)

def kpis(df: pd.DataFrame) -> dict:
    s1 = pd.Series([_savings_for_kpi(x) for x in df.get('Alt1 Savings %', [])], dtype=float)
    s1 = s1.dropna()
    return {'total': len(df), 'avg1': s1.mean() if len(s1) else None, 'max1': s1.max() if len(s1) else None, 'act_col': 'Actual Cost ($)' in df.columns}

def render_kpis(k: dict):
    (c1, c2, c3, c4) = st.columns(4)
    with c1:
        st.markdown(f"""<div class="metric"><div class="metric-label">Rows</div><div class="metric-value">{k['total']:,}</div></div>""", unsafe_allow_html=True)
    with c2:
        v = f"{k['avg1']:.1f}%" if k['avg1'] is not None else '—'
        st.markdown(f'<div class="metric"><div class="metric-label">Avg Alt1 savings</div><div class="metric-value">{v}</div></div>', unsafe_allow_html=True)
    with c3:
        v = f"{k['max1']:.1f}%" if k['max1'] is not None else '—'
        st.markdown(f'<div class="metric"><div class="metric-label">Max Alt1 savings</div><div class="metric-value">{v}</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="metric"><div class="metric-label">Actual cost column</div><div class="metric-value">{('Yes' if k['act_col'] else 'No')}</div></div>""", unsafe_allow_html=True)
for (key, default) in (('load_result', None), ('result', None), ('region_id', DEFAULT_REGION), ('service', 'both'), ('cpu_filter', 'both'), ('cost_pick', None)):
    if key not in st.session_state:
        st.session_state[key] = default
stale = cache_is_stale()
pill = f'Cache {cache_age_days()}d — refresh due' if stale else 'Cache fresh'
st.markdown(f'\n<div class="hdr">\n  <h1>💰 FinOps Optimizer <span style="opacity:0.85;font-size:0.75rem">({pill})</span></h1>\n  <p>EC2 & RDS · Actual-cost savings · Local pricing only · Original columns preserved</p>\n</div>\n', unsafe_allow_html=True)
(up_col, reg_col, svc_col, cpu_col, go_col) = st.columns([3, 2, 1, 1, 1])
with up_col:
    uploaded = st.file_uploader('file', type=['csv', 'xlsx', 'xls'], label_visibility='collapsed')
    st.caption('Upload CSV / Excel — all columns kept in order')
with reg_col:
    st.markdown('<div class="sec">Pricing region</div>', unsafe_allow_html=True)
    region_opts = [f'{label}  [{rid}]' for (rid, label) in SUPPORTED_REGIONS]
    default_idx = [r for (r, _) in SUPPORTED_REGIONS].index(DEFAULT_REGION)
    sel_disp = st.selectbox('region', region_opts, index=default_idx, label_visibility='collapsed')
    sel_region = [r for (r, _) in SUPPORTED_REGIONS][region_opts.index(sel_disp)]
    st.session_state['region_id'] = sel_region
with svc_col:
    st.markdown('<div class="sec">Service</div>', unsafe_allow_html=True)
    st.session_state['service'] = st.radio('svc', ['ec2', 'rds', 'both'], format_func=lambda x: {'ec2': 'EC2', 'rds': 'RDS', 'both': 'Both'}[x], label_visibility='collapsed', horizontal=True)
with cpu_col:
    st.markdown('<div class="sec">CPU</div>', unsafe_allow_html=True)
    st.session_state['cpu_filter'] = st.selectbox('cpu', ['both', 'default', 'intel', 'graviton'], format_func=lambda x: {'both': 'Both', 'default': 'Default', 'intel': 'Intel', 'graviton': 'Graviton'}[x], label_visibility='collapsed')
with go_col:
    st.markdown('<br>', unsafe_allow_html=True)
    run = st.button('Process', type='primary', disabled=uploaded is None, use_container_width=True)
st.markdown('---')
if run and uploaded:
    with st.spinner('Loading…'):
        try:
            lr: LoadResult = load_file(uploaded, uploaded.name)
            st.session_state['load_result'] = lr
            st.session_state['result'] = None
            st.session_state['cost_pick'] = None
            st.session_state['binding'] = None
            st.session_state.pop('_enrich_svc', None)
            st.session_state.pop('_enrich_cpu', None)
            for w in lr.warnings:
                st.markdown(f'<div class="box-warn">⚠️ {w}</div>', unsafe_allow_html=True)
        except ValueError as ve:
            st.session_state['load_result'] = None
            st.markdown(f'<div class="box-err">❌ {ve}</div>', unsafe_allow_html=True)
        except Exception as e:
            st.session_state['load_result'] = None
            log.error('load_file failed: %s', type(e).__name__)
            st.markdown('<div class="box-err">❌ Failed to read file.</div>', unsafe_allow_html=True)
lr: LoadResult | None = st.session_state.get('load_result')
binding_ready = False
chosen_binding = None
if st.session_state.get('binding') is not None:
    chosen_binding = st.session_state['binding']
    binding_ready = True
if lr is not None and (not binding_ready):
    cols_all = list(lr.df.columns)
    if lr.needs_instance_pick or lr.needs_os_pick:
        st.markdown('<div class="sec">Column mapping (required)</div>', unsafe_allow_html=True)
        (mc1, mc2) = st.columns(2)
        with mc1:
            di = 0
            if lr.instance_candidates:
                di = cols_all.index(lr.instance_candidates[0]) if lr.instance_candidates[0] in cols_all else 0
            inst_sel = st.selectbox('Instance / DB class column', cols_all, index=min(di, len(cols_all) - 1))
        with mc2:
            do = 0
            if lr.os_candidates:
                do = cols_all.index(lr.os_candidates[0]) if lr.os_candidates[0] in cols_all else 0
            os_sel = st.selectbox('OS / engine column', cols_all, index=min(do, len(cols_all) - 1))
        cost_sel = None
        if len(lr.cost_candidates) > 1:
            cost_sel = st.selectbox('Actual cost column (required for savings)', lr.cost_candidates, key='cost_ambiguous')
        elif len(lr.cost_candidates) == 1:
            cost_sel = lr.cost_candidates[0]
        else:
            cost_sel = st.selectbox('Actual cost column (optional)', ['— None —'] + cols_all, key='cost_optional')
            if cost_sel == '— None —':
                cost_sel = None
        if st.button('Apply column mapping', type='primary'):
            try:
                b = finalize_binding(lr, inst_sel, os_sel, cost_sel).binding
                st.session_state['binding'] = b
                st.session_state['cost_pick'] = cost_sel
                st.rerun()
            except ValueError as e:
                st.markdown(f'<div class="box-err">❌ {e}</div>', unsafe_allow_html=True)
    elif lr.binding is not None:
        if len(lr.cost_candidates) > 1 and lr.binding.actual_cost is None:
            st.markdown('<div class="box-warn">Multiple cost columns — pick one for Actual Cost.</div>', unsafe_allow_html=True)
            cp = st.selectbox('Actual cost column', lr.cost_candidates, key='cost_pick_multi')
            if st.button('Confirm cost column', type='primary'):
                b = finalize_binding(lr, lr.binding.instance, lr.binding.os, cp).binding
                st.session_state['binding'] = b
                st.rerun()
        elif st.session_state.get('binding') is None:
            st.session_state['binding'] = lr.binding
if lr is not None and st.session_state.get('binding') is not None:
    chosen_binding = st.session_state['binding']
    binding_ready = True
    if st.session_state.get('result') is None:
        if st.button('Run enrichment', type='primary', key='run_enrich'):
            try:
                svc = st.session_state['service']
                cpu = st.session_state['cpu_filter']
                reg = st.session_state.get('region_id', DEFAULT_REGION)
                out = process(lr.df, chosen_binding, region=reg, service=svc, cpu_filter=cpu)
                st.session_state['result'] = out
                st.session_state['_enrich_svc'] = svc
                st.session_state['_enrich_cpu'] = cpu
                st.success(f'Enriched {len(out):,} rows')
                st.rerun()
            except Exception as ex:
                st.markdown(f'<div class="box-err">❌ {ex}</div>', unsafe_allow_html=True)
                log.error('process failed: %s', type(ex).__name__)
df_out: pd.DataFrame | None = st.session_state.get('result')
if df_out is not None:
    if st.session_state.get('_enrich_svc') != st.session_state.get('service') or st.session_state.get('_enrich_cpu') != st.session_state.get('cpu_filter'):
        st.warning('Service or CPU mode changed since last enrichment — click **Run enrichment** to refresh.')
    st.markdown('<div class="sec">Filter bar</div>', unsafe_allow_html=True)
    (f1, f2, f3, f4) = st.columns([1, 1, 1, 3])
    with f1:
        vf_svc = st.radio('View service', ['all', 'ec2', 'rds'], format_func=lambda x: {'all': 'Both (show all)', 'ec2': 'EC2 rows only', 'rds': 'RDS rows only'}[x], horizontal=True)
    with f2:
        st.caption('CPU (enrichment)')
        st.write(str(st.session_state.get('cpu_filter', 'both')).title())
    with f3:
        vf_os = st.text_input('OS contains', placeholder='filter…')
    with f4:
        q = st.text_input('Search', placeholder='any column…')
    view = df_out.copy()
    bind = st.session_state.get('binding')
    inst_col = bind.instance if bind else None
    if vf_svc == 'ec2' and inst_col and (inst_col in view.columns):
        view = view[~view[inst_col].astype(str).str.lower().str.strip().str.startswith('db.')]
    elif vf_svc == 'rds' and inst_col and (inst_col in view.columns):
        view = view[view[inst_col].astype(str).str.lower().str.strip().str.startswith('db.')]
    os_col_name = bind.os if bind else None
    if vf_os and os_col_name and (os_col_name in view.columns):
        view = view[view[os_col_name].astype(str).str.contains(vf_os, case=False, na=False)]
    if q:
        m = pd.Series(False, index=view.index)
        for col in view.columns:
            m |= view[col].astype(str).str.contains(q, case=False, na=False)
        view = view[m]
    st.caption(f'Showing **{len(view):,}** of **{len(df_out):,}** rows')
    render_kpis(kpis(view))

    def _style_sav(v: str) -> str:
        n = _savings_numeric(v)
        if n is None:
            return 'color: #9ca3af'
        if n >= 20:
            return 'color: #22c55e; font-weight:700'
        if n > 0:
            return 'color: #d97706; font-weight:600'
        return 'color: #dc2626; font-weight:600'

    def _fmt_cell(cname: str):

        def _f(x):
            if 'Savings %' in cname:
                if pd.isna(x) or x is None:
                    return 'N/A'
                return str(x) if isinstance(x, str) else f'{float(x):.1f}%'
            if 'Cost ($)' in cname:
                if pd.isna(x) or x is None:
                    return 'N/A'
                if isinstance(x, (int, float)):
                    return f'${float(x):,.4f}'
                return str(x)
            return x
        return _f
    fmt_map = {c: _fmt_cell(c) for c in view.columns if 'Savings %' in c or 'Cost ($)' in c}
    style_cols = {c: _style_sav for c in view.columns if 'Savings %' in c}
    try:
        sty = view.style.format(fmt_map, na_rep='N/A')
        for (col_name, fn) in style_cols.items():
            sty = sty.map(fn, subset=[col_name])
        sty = sty.set_properties(**{'font-size': '0.8rem'})
    except Exception:
        sty = view
    st.dataframe(sty, use_container_width=True, hide_index=True, height=480)
    export_df = apply_na_fill(df_out)
    reg_lbl = REGION_LABELS.get(st.session_state.get('region_id', DEFAULT_REGION), '')
    st.download_button('Download Excel', build_excel(export_df, reg_lbl), 'finops_recommendations.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    st.download_button('Download CSV', export_df.to_csv(index=False).encode(), 'finops_recommendations.csv', 'text/csv')
elif lr is None:
    st.info('Upload a file and click **Process** to load. Then map columns if prompted, **Run enrichment**.')
elif not binding_ready:
    st.info('Complete **column mapping** (and cost selection if needed), then **Run enrichment**.')
