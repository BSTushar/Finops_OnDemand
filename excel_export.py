from __future__ import annotations
import io
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pricing_engine import CACHE_METADATA, DECISION_SUPPORT_NOTE, PRICING_SOURCE_LABEL, cost_disclaimer_text, format_pricing_snapshot_line


def savings_numeric(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        if v.strip() == 'No Savings':
            return 0.0
        try:
            return float(v.replace('%', ''))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_excel(df: pd.DataFrame, region_label: str, pricing_region_id: str) -> bytes:
    buf = io.BytesIO()
    preamble_rows = 3
    startrow = preamble_rows + 1
    ncol = max(len(df.columns), 1)
    end_letter = get_column_letter(ncol)
    snapshot = format_pricing_snapshot_line(pricing_region_id)
    disclaimer = cost_disclaimer_text(pricing_region_id)
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Recommendations', startrow=startrow)
        wb = writer.book
        ws = writer.sheets['Recommendations']
        thin = Side(style='thin', color='888888')
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
        disc_font = Font(size=9, bold=True)
        note_font = Font(size=9)
        disc_fill = PatternFill('solid', fgColor='FEF3C7')
        for (r, txt, font, fill) in (
            (1, disclaimer, disc_font, disc_fill),
            (2, snapshot, note_font, None),
            (3, DECISION_SUPPORT_NOTE, note_font, None),
        ):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            c = ws.cell(row=r, column=1)
            c.value = txt
            c.font = font
            if fill is not None:
                c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical='center')
            c.border = bdr
        hdr_row = startrow + 1
        hdr_fill = PatternFill('solid', fgColor='E5E7EB')
        hdr_font = Font(bold=True, color='111827', size=10)
        for cidx in range(1, ncol + 1):
            c = ws.cell(row=hdr_row, column=cidx)
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sav_cols = [c for c in df.columns if 'Savings %' in c]
        price_cols = [c for c in df.columns if 'Cost ($)' in c]
        green_fill = PatternFill('solid', fgColor='D1FAE5')
        amber_fill = PatternFill('solid', fgColor='FEF3C7')
        red_fill = PatternFill('solid', fgColor='FEE2E2')
        col_list = list(df.columns)
        first_data_row = hdr_row + 1
        for ridx in range(first_data_row, ws.max_row + 1):
            for cidx in range(1, ncol + 1):
                cell = ws.cell(row=ridx, column=cidx)
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(size=10)
            for name in sav_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                sc = ws.cell(row=ridx, column=ci)
                val = sc.value
                nv = savings_numeric(val)
                if nv is None:
                    continue
                if nv >= 20:
                    sc.fill = green_fill
                elif nv > 0:
                    sc.fill = amber_fill
                else:
                    sc.fill = red_fill
            for name in price_cols:
                if name not in col_list:
                    continue
                ci = col_list.index(name) + 1
                ws.cell(row=ridx, column=ci).number_format = '$#,##0.0000'
        for cidx in range(1, ncol + 1):
            letter = get_column_letter(cidx)
            w = max((len(str(ws.cell(row=r, column=cidx).value or '')) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[letter].width = min(w + 2, 48)
        ws.freeze_panes = ws.cell(row=first_data_row, column=1).coordinate
        ws.auto_filter.ref = f'A{hdr_row}:{end_letter}{ws.max_row}'
        ws_m = wb.create_sheet('Metadata')
        ws_m.append(['Field', 'Value'])
        ws_m.append(['Disclaimer', disclaimer])
        ws_m.append(['Pricing snapshot', snapshot])
        ws_m.append(['Decision support note', DECISION_SUPPORT_NOTE])
        ws_m.append(['Generated at', datetime.now().strftime('%Y-%m-%d %H:%M')])
        ws_m.append(['Pricing region (label)', region_label])
        ws_m.append(['Pricing region (id)', pricing_region_id])
        ws_m.append(['Pricing source', PRICING_SOURCE_LABEL])
        ws_m.append(['Dataset as-of', CACHE_METADATA['last_updated'].strftime('%Y-%m-%d')])
        ws_m.append(['Rows (data)', len(df)])
    return buf.getvalue()
